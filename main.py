from fastapi import FastAPI, Request
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
import requests as http_client
from datetime import date, datetime
from urllib.parse import quote
import asyncio
import json as _json
import io

app = FastAPI()
templates = Jinja2Templates(directory="templates")

checkin_set: set[str] = set()
_sse_clients: set[asyncio.Queue] = set()


async def _broadcast(event: str, data: dict) -> None:
    """연결된 모든 SSE 클라이언트에 이벤트 전송"""
    if not _sse_clients:
        return
    payload = _json.dumps(data, ensure_ascii=False)
    dead: set = set()
    for q in _sse_clients:
        try:
            q.put_nowait((event, payload))
        except Exception:
            dead.add(q)
    _sse_clients.difference_update(dead)

BASE_URL = "https://cde.jj.ac.kr/_custom/jj/_common/app/room-reservation/logic/ajax.jsp"
WEEKDAY_KO = ["월", "화", "수", "목", "금", "토", "일"]


def _fetch(d: str) -> list | None:
    try:
        r = http_client.get(BASE_URL, params={"mode": "day-list", "date": d}, timeout=10)
        return r.json().get("items", [])
    except Exception:
        return None


def _build_context() -> dict:
    today = date.today()
    d = today.strftime("%Y-%m-%d")
    items = _fetch(d)
    sorted_items = sorted(items or [], key=lambda x: x.get("rrStartTime", ""))
    return {
        "date_label": f"{today.year}년 {today.month}월 {today.day}일 ({WEEKDAY_KO[today.weekday()]})",
        "today_ym": today.strftime("%Y-%m"),
        "items": sorted_items,
        "checkin_list": list(checkin_set),
        "checkin_set": checkin_set,
        "error": items is None,
        "updated_at": datetime.now().strftime("%H:%M"),
        "total": len(sorted_items),
        "confirmed": sum(1 for i in sorted_items if i.get("rrState") == "예약완료"),
    }


@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return templates.TemplateResponse(request=request, name="index.html", context=_build_context())


@app.get("/table", response_class=HTMLResponse)
async def table_partial(request: Request):
    return templates.TemplateResponse(request=request, name="_table.html", context=_build_context())


@app.get("/events")
async def sse_stream(request: Request):
    """SSE 스트림 — 입실 변경 시 다른 기기에 실시간 이벤트 전송"""
    async def generator():
        q: asyncio.Queue = asyncio.Queue(maxsize=20)
        _sse_clients.add(q)
        try:
            yield "data: ok\n\n"
            while True:
                if await request.is_disconnected():
                    break
                try:
                    event, payload = await asyncio.wait_for(q.get(), timeout=20)
                    yield f"event: {event}\ndata: {payload}\n\n"
                except asyncio.TimeoutError:
                    yield ": keep-alive\n\n"
        finally:
            _sse_clients.discard(q)

    return StreamingResponse(
        generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        },
    )


@app.post("/checkin/{rr_seq}", response_class=HTMLResponse)
async def toggle_checkin(request: Request, rr_seq: str):
    tab = request.headers.get("X-Tab-Id", "")
    action = "out" if rr_seq in checkin_set else "in"
    if rr_seq in checkin_set:
        checkin_set.discard(rr_seq)
    else:
        checkin_set.add(rr_seq)

    ctx = _build_context()
    item = next((i for i in ctx["items"] if str(i.get("rrSeq", "")) == rr_seq), {})

    await _broadcast("checkin_changed", {
        "tab": tab,
        "action": action,
        "rr_seq": rr_seq,
        "rpName": item.get("rpName", ""),
        "rrBooker": item.get("rrBooker", ""),
        "rrStartTime": item.get("rrStartTime", ""),
    })

    return templates.TemplateResponse(request=request, name="_table.html", context=ctx)


def _fetch_month(y: int, m: int) -> dict[int, list]:
    """해당 월의 모든 날 병렬 조회 → {day: [items]} (예약 있는 날만)"""
    import calendar
    from concurrent.futures import ThreadPoolExecutor, as_completed

    _, last_day = calendar.monthrange(y, m)
    day_strs = [f"{y:04d}-{m:02d}-{day:02d}" for day in range(1, last_day + 1)]
    day_data: dict[int, list] = {}
    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(_fetch, d): int(d.split("-")[2]) for d in day_strs}
        for future in as_completed(futures):
            day = futures[future]
            items = future.result()
            if items:
                day_data[day] = sorted(items, key=lambda x: x.get("rrStartTime", ""))
    return day_data


@app.get("/export")
async def export_excel(year: int | None = None, month: int | None = None):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "openpyxl not installed"}, status_code=500)

    import calendar

    today = date.today()
    y = year or today.year
    m = month or today.month
    _, last_day = calendar.monthrange(y, m)

    day_data = _fetch_month(y, m)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{y}년 {m}월"

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 50

    COL_C_CHARS = 50  # C열 너비(글자 수) — 행 높이 계산용

    DATE_FILL  = PatternFill("solid", fgColor="1E3A8A")
    DATE_FONT  = Font(bold=True, color="FFFFFF", size=11)
    COL_FILL   = PatternFill("solid", fgColor="DBEAFE")
    COL_FONT   = Font(bold=True, color="1E40AF", size=10)
    THIN       = Side(style="thin", color="CBD5E1")
    BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    cur = 1

    for day in range(1, last_day + 1):
        if day not in day_data:
            continue

        d   = date(y, m, day)
        lbl = f"{y}년 {m}월 {day}일 ({WEEKDAY_KO[d.weekday()]})"

        # 날짜 헤더
        ws.merge_cells(f"A{cur}:C{cur}")
        cell = ws.cell(row=cur, column=1, value=lbl)
        cell.font      = DATE_FONT
        cell.fill      = DATE_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[cur].height = 20
        cur += 1

        # 컬럼 헤더
        for col, header in enumerate(["스튜디오", "이름", "사유"], 1):
            cell = ws.cell(row=cur, column=col, value=header)
            cell.font      = COL_FONT
            cell.fill      = COL_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = BORDER
        cur += 1

        # 데이터 행
        for item in day_data[day]:
            purpose = item.get("rrPurpose", "")
            # 단락별로 줄 수 계산 (한글 2칸, 영문 1칸)
            total_lines = 0
            for para in (purpose.split("\n") if purpose else [""]):
                cw = sum(2 if ord(c) > 127 else 1 for c in para)
                total_lines += max(1, -(-cw // COL_C_CHARS))
            ws.row_dimensions[cur].height = 15 * (total_lines + 1)

            aligns = ["left", "center", "left"]
            for col, (val, align) in enumerate(zip(
                [item.get("rpName", ""), item.get("rrBooker", ""), purpose],
                aligns,
            ), 1):
                cell = ws.cell(row=cur, column=col, value=val)
                cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=(col == 3))
                cell.border    = BORDER
            cur += 1

        cur += 1  # 날짜 사이 빈 행

    if cur == 1:
        ws.cell(row=1, column=1, value=f"{y}년 {m}월 예약 내역 없음")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"CDE스튜디오_{y}년{m:02d}월.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@app.get("/export/today")
async def export_today_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "openpyxl not installed"}, status_code=500)

    today = date.today()
    d = today.strftime("%Y-%m-%d")
    items = _fetch(d) or []
    items = sorted(items, key=lambda x: x.get("rrStartTime", ""))
    lbl = f"{today.year}년 {today.month}월 {today.day}일 ({WEEKDAY_KO[today.weekday()]})"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = lbl

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14

    THIN   = Side(style="thin", color="CBD5E1")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    ws.merge_cells("A1:B1")
    c = ws.cell(row=1, column=1, value=lbl)
    c.font      = Font(bold=True, color="FFFFFF", size=11)
    c.fill      = PatternFill("solid", fgColor="1E3A8A")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 20

    for col, header in enumerate(["스튜디오", "예약자"], 1):
        c = ws.cell(row=2, column=col, value=header)
        c.font      = Font(bold=True, color="1E40AF", size=10)
        c.fill      = PatternFill("solid", fgColor="DBEAFE")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = BORDER

    for r, item in enumerate(items, 3):
        for col, val in enumerate([item.get("rpName", ""), item.get("rrBooker", "")], 1):
            c = ws.cell(row=r, column=col, value=val)
            c.alignment = Alignment(horizontal="left" if col == 1 else "center", vertical="center")
            c.border    = BORDER

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"CDE스튜디오_{d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@app.get("/export/today/txt")
async def export_today_txt():
    today = date.today()
    d = today.strftime("%Y-%m-%d")
    items = _fetch(d) or []
    items = sorted(items, key=lambda x: x.get("rrStartTime", ""))
    lbl = f"{today.year}년 {today.month}월 {today.day}일 ({WEEKDAY_KO[today.weekday()]})"

    lines = [lbl, "─" * 30]
    for item in items:
        lines.append(f"{item.get('rpName', '')} | {item.get('rrBooker', '')}")
    lines.append("")

    buf = io.BytesIO("\n".join(lines).encode("utf-8-sig"))
    filename = f"CDE스튜디오_{d}.txt"
    return StreamingResponse(
        buf,
        media_type="text/plain; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@app.get("/export/txt")
async def export_txt(year: int | None = None, month: int | None = None):
    import calendar

    today = date.today()
    y = year or today.year
    m = month or today.month
    _, last_day = calendar.monthrange(y, m)

    day_data = _fetch_month(y, m)

    lines: list[str] = [f"■ {y}년 {m}월 CDE 스튜디오 예약 현황\n"]

    for day in range(1, last_day + 1):
        if day not in day_data:
            continue

        d   = date(y, m, day)
        lbl = f"{y}년 {m}월 {day}일 ({WEEKDAY_KO[d.weekday()]})"
        lines.append(f"{'─' * 40}")
        lines.append(lbl)
        lines.append(f"{'─' * 40}")

        for item in day_data[day]:
            studio  = item.get("rpName",    "")
            booker  = item.get("rrBooker",  "")
            purpose = item.get("rrPurpose", "")
            lines.append(f"{studio} | {booker} | {purpose}")

        lines.append("")

    content = "\n".join(lines)
    buf = io.BytesIO(content.encode("utf-8-sig"))  # BOM 포함 → 메모장 한글 깨짐 방지

    filename = f"CDE스튜디오_{y}년{m:02d}월.txt"
    return StreamingResponse(
        buf,
        media_type="text/plain; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )
